"""Weekend blackout window + reserved-item exclusion (manifest_url form).

Covers:
- The Problem A fix: reserved (immutable/promo) items are excluded from public
  results even when ``saved_playlist_items.media_id`` stores the MANIFEST URL
  (production form), not the bare friendly_token.
- The blackout table: blacked-out items are hidden from regular users, surfaced
  (with ``blackout_active``) for admins, and released once expired.
- The workbook scan + expiry math (blackout lifts after the *following* weekend).
"""

import datetime as _dt
import io

import pytest

from kryten_webqueue.catalog.db import Database
from kryten_webqueue.integrations.cmsutils.blackout_scan import (
    _expiry_for_friday,
    _sheet_weekend_dates,
    scan_workbook_blackouts,
)

MEDIACMS = "https://www.dropsugar.co"


def _manifest(token: str) -> str:
    return f"{MEDIACMS}/api/v1/media/cytube/{token}.json?format=json"


@pytest.fixture
async def db(tmp_path):
    database = Database(str(tmp_path / "blackout.db"))
    await database.connect()
    await database.run_migrations()
    yield database
    await database.close()


async def _add_catalog(db, token, title):
    await db.insert_catalog(
        {
            "friendly_token": token,
            "title": title,
            "description": "",
            "duration_sec": 600,
            "manifest_url": _manifest(token),
            "thumbnail_url": "",
            "synced_at": "2026-01-01T00:00:00+00:00",
        }
    )


# ── Problem A: reserved exclusion with manifest_url-form media_id ──────────────


async def test_reserved_excluded_when_media_id_is_manifest_url(db):
    """Regression: production stores media_id as the manifest URL, not the token."""
    await _add_catalog(db, "keeptok", "Keep Me Visible")
    await _add_catalog(db, "restrtok", "Reserved Weekend Film")

    pl_id = await db.create_saved_playlist(
        name="Friday Night",
        description=None,
        is_immutable=True,
        created_by="admin",
    )
    # Stored exactly as the importer does: the manifest URL.
    await db.append_playlist_item(
        pl_id,
        {
            "media_type": "cm",
            "media_id": _manifest("restrtok"),
            "title": "Reserved Weekend Film",
            "duration_sec": 600,
        },
    )

    browse_tokens = {r["friendly_token"] for r in await db.browse()}
    assert "keeptok" in browse_tokens
    assert "restrtok" not in browse_tokens
    assert await db.browse_count() == 1
    assert (await db.get_item("restrtok")) is None
    assert await db.is_restricted("restrtok") is True
    assert await db.is_restricted("keeptok") is False


# ── Blackout table behaviour ──────────────────────────────────────────────────


async def test_blackout_hides_from_users_but_admin_sees_flag(db):
    await _add_catalog(db, "bo1", "Falling Down")
    await _add_catalog(db, "normal1", "Ordinary Movie")

    future = (_dt.datetime.now(_dt.UTC) + _dt.timedelta(days=9)).isoformat(sep=" ")
    await db.upsert_blackout("bo1", reason="weekend:9.5-9.6", expires_at=future)

    # Regular user: blacked-out item hidden from browse/search/get_item.
    user_tokens = {r["friendly_token"] for r in await db.browse()}
    assert "bo1" not in user_tokens
    assert "normal1" in user_tokens
    assert await db.browse_count() == 1
    assert (await db.get_item("bo1")) is None
    assert await db.is_blackout("bo1") is True

    # Admin (show_hidden): item present and flagged.
    admin_rows = {r["friendly_token"]: r for r in await db.browse(show_hidden=True)}
    assert "bo1" in admin_rows
    assert admin_rows["bo1"]["blackout_active"] == 1
    assert admin_rows["normal1"]["blackout_active"] == 0


async def test_expired_blackout_does_not_hide(db):
    await _add_catalog(db, "old1", "Last Month's Feature")
    past = (_dt.datetime.now(_dt.UTC) - _dt.timedelta(days=1)).isoformat(sep=" ")
    await db.upsert_blackout("old1", reason="weekend:1.1-1.2", expires_at=past)

    assert await db.is_blackout("old1") is False
    assert "old1" in {r["friendly_token"] for r in await db.browse()}
    assert (await db.get_item("old1")) is not None

    removed = await db.prune_expired_blackouts()
    assert removed == 1
    assert await db.count_active_blackouts() == 0


async def test_upsert_keeps_latest_expiry(db):
    await _add_catalog(db, "ext1", "Held Over")
    future_early = (_dt.datetime.now(_dt.UTC) + _dt.timedelta(days=7)).isoformat(
        sep=" "
    )
    future_late = (_dt.datetime.now(_dt.UTC) + _dt.timedelta(days=21)).isoformat(
        sep=" "
    )
    await db.upsert_blackout("ext1", reason="weekend:9.5-9.6", expires_at=future_late)
    # A later re-scan reporting an earlier expiry must not shorten the window.
    await db.upsert_blackout("ext1", reason="weekend:9.5-9.6", expires_at=future_early)
    stored = {
        r["friendly_token"]: r["expires_at"] for r in await db.list_active_blackouts()
    }
    assert stored.get("ext1") == future_late


# ── Scan + expiry math ────────────────────────────────────────────────────────


def test_sheet_weekend_dates_and_expiry():
    today = _dt.date(2026, 3, 4)  # Wednesday
    friday, sunday = _sheet_weekend_dates("3.6-3.7", today)
    assert friday == _dt.date(2026, 3, 6)
    assert sunday == _dt.date(2026, 3, 8)
    # Blackout lifts the Monday after the FOLLOWING weekend (Friday + 10).
    assert _expiry_for_friday(friday) == "2026-03-16 00:00:00"
    assert _sheet_weekend_dates("not a sheet", today) is None


def test_scan_workbook_collects_future_sheet_dropsugar_urls():
    import openpyxl

    today = _dt.date(2026, 3, 4)
    wb = openpyxl.Workbook()

    # Future/current weekend sheet — should be scanned.
    ws = wb.active
    ws.title = "3.6-3.7"
    ws.cell(row=1, column=5, value="https://www.dropsugar.co/view?m=TOKENE")  # col E
    ws.cell(row=1, column=6, value=_manifest("TOKENF"))  # col F
    ws.cell(
        row=2, column=5, value="https://www.youtube.com/watch?v=xxxxxxxxxxx"
    )  # skip

    # Past weekend sheet — should be skipped.
    past = wb.create_sheet("2.27-2.28")
    past.cell(row=1, column=6, value=_manifest("OLDTOK"))

    # Non-date sheet — ignored.
    wb.create_sheet("Played Movies").cell(row=1, column=6, value=_manifest("NOPE"))

    buf = io.BytesIO()
    wb.save(buf)

    blackouts, sheets_scanned = scan_workbook_blackouts(buf.getvalue(), today)
    tokens = {b["token"]: b["expires_at"] for b in blackouts}
    assert sheets_scanned == 1
    assert tokens == {
        "TOKENE": "2026-03-16 00:00:00",
        "TOKENF": "2026-03-16 00:00:00",
    }
