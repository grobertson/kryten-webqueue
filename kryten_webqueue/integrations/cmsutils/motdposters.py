"""MOTD poster generator.

Fetches OMDB poster art for every movie in the upcoming weekend sheet,
saves images to the configured motd_boxes directory, and writes a
``<div class="poster-grid">`` HTML snippet ready to paste into CyTube's
Message of the Day.

Poster filename format:  art-YYYY-MM-DD-movieN-nightM.jpg
  N  = 1-based index within that night (across all sections for that day)
  M  = 1 (Friday), 2 (Saturday), 3 (Sunday)

OMDB search: GET https://www.omdbapi.com/?t={title}&type=movie&y={year}&apikey=…
Poster image: GET https://img.omdbapi.com/?i={imdbID}&h=600&apikey=…
"""

from __future__ import annotations

import datetime
import io
import logging
import random
import re
import time
from html import escape
from pathlib import Path

import requests

from .fetchurls import (
    _SHEET_DATE_RE,
    acquire_graph_token_silent,
    download_sharepoint_xlsx,
    upcoming_weekend_sheet,
)

logger = logging.getLogger(__name__)

OMDB_SEARCH_URL = "https://www.omdbapi.com/"
OMDB_IMAGE_URL = "https://img.omdbapi.com/"
REQUEST_TIMEOUT = 15
BETWEEN_REQUESTS_DELAY = 0.3  # polite gap between OMDB calls

# All section header keywords, checked in order (more specific first)
_SECTION_MAP = {
    "friday": "friday",
    "saturday morning": "saturday-morning",
    "saturday": "saturday-night",
    "sunday morning": "sunday-morning",
    "sunday afternoon": "sunday-daytime",
}

# Section slug → night number (1=Friday, 2=Saturday, 3=Sunday)
_NIGHT_MAP: dict[str, int] = {
    "friday": 1,
    "saturday-morning": 2,
    "saturday-night": 2,
    "sunday-morning": 3,
    "sunday-daytime": 3,
}

_NIGHT_LABELS: dict[int, str] = {
    1: "Friday Night",
    2: "Saturday Night",
    3: "Sunday Night",
}

# Canonical section processing order (determines night/movie numbering)
_SECTION_ORDER = [
    "friday",
    "saturday-morning",
    "saturday-night",
    "sunday-morning",
    "sunday-daytime",
]

_IMAGE_EXTS = {".webp", ".jpg", ".jpeg", ".png", ".gif", ".avif"}


# ── Sheet parsing ──────────────────────────────────────────────────────────────


def _classify_section(cell_value: str) -> str | None:
    if not cell_value:
        return None
    v = cell_value.strip().lower()
    for keyword, slug in _SECTION_MAP.items():
        if keyword in v:
            return slug
    return None


def _is_movie_tag(value: str) -> bool:
    v = value.strip().lower()
    return v == "movie" or v == "movie event"


def extract_movies_by_section(wb_bytes: bytes, sheet_name: str) -> dict[str, list[str]]:
    """Return {section_slug: [title, ...]} for every movie row in the sheet."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for motd_posters")

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]
    current_section: str | None = None
    movies: dict[str, list[str]] = {}

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0]).strip() if row[0] is not None else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        slug = _classify_section(col_a)
        if slug:
            current_section = slug
            continue

        if current_section and _is_movie_tag(col_b) and col_c:
            movies.setdefault(current_section, []).append(col_c)

    wb.close()
    return movies


# ── OMDB helpers ───────────────────────────────────────────────────────────────


def _extract_year(title: str) -> str | None:
    m = re.search(r"\b((?:19|20)\d{2})\b", title)
    return m.group(1) if m else None


def _strip_year(title: str, year: str) -> str:
    cleaned = re.sub(r"\s*[\(\[]?" + re.escape(year) + r"[\)\]]?", "", title)
    return cleaned.strip(" .-") or title


def _omdb_lookup(title: str, *, api_key: str, year: str | None = None) -> dict | None:
    """Return {"imdb_id": str, "poster_url": str} or None."""
    params: dict[str, str] = {"apikey": api_key, "t": title, "type": "movie"}
    if year:
        params["y"] = year
    try:
        resp = requests.get(OMDB_SEARCH_URL, params=params, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            logger.warning("OMDB HTTP %s for %r", resp.status_code, title)
            return None
        data = resp.json()
        if data.get("Response") == "False":
            return None
        imdb_id = data.get("imdbID") or ""
        if not imdb_id:
            return None
        poster_url = f"{OMDB_IMAGE_URL}?i={imdb_id}&h=600&apikey={api_key}"
        return {"imdb_id": imdb_id, "poster_url": poster_url}
    except requests.RequestException as exc:
        logger.warning("OMDB request error for %r: %s", title, exc)
        return None


def _resolve_omdb(title: str, *, api_key: str) -> dict | None:
    """Try OMDB with year (if in title) then without, accepting the first hit."""
    year = _extract_year(title)
    clean = _strip_year(title, year) if year else title

    # Try cleaned-with-year first, then original-with-year, then clean-no-year
    attempts: list[tuple[str, str | None]] = []
    if year:
        attempts.append((clean, year))
        if clean != title:
            attempts.append((title, year))
    attempts.append((clean, None))
    if clean != title:
        attempts.append((title, None))

    seen: set[tuple] = set()
    for t, y in attempts:
        key = (t, y)
        if key in seen:
            continue
        seen.add(key)
        result = _omdb_lookup(t, api_key=api_key, year=y)
        if result:
            return result
        time.sleep(BETWEEN_REQUESTS_DELAY)

    return None


# ── Image I/O ──────────────────────────────────────────────────────────────────


def _download_image(url: str) -> bytes | None:
    try:
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200 and resp.content:
            return resp.content
        logger.warning("Image download failed %s: HTTP %s", url, resp.status_code)
    except requests.RequestException as exc:
        logger.warning("Image download error %s: %s", url, exc)
    return None


def _pick_placeholder(placeholder_dir: str) -> bytes | None:
    p = Path(placeholder_dir).expanduser()
    if not p.is_dir():
        return None
    candidates = [f for f in p.iterdir() if f.is_file() and f.suffix.lower() in _IMAGE_EXTS]
    if not candidates:
        return None
    return random.choice(candidates).read_bytes()


# ── Filename / HTML generation ─────────────────────────────────────────────────


def _poster_filename(date: datetime.date, movie_idx: int, night: int) -> str:
    return f"art-{date:%Y-%m-%d}-movie{movie_idx}-night{night}.jpg"


def _generate_html(night_entries: list[dict], poster_base_url: str) -> str:
    """Build the <div class="poster-grid"> snippet from processed entries."""
    base = poster_base_url.rstrip("/")
    lines = ['<div class="poster-grid">']
    current_night: int | None = None

    for entry in night_entries:
        night = entry["night"]
        if night != current_night:
            if current_night is not None:
                lines.append("")
            lines.append(f"    <!-- {_NIGHT_LABELS.get(night, f'Night {night}')} -->")
            current_night = night

        img_url = f"{base}/{entry['filename']}"
        imdb_id = entry.get("imdb_id") or ""
        href = f"https://www.imdb.com/title/{imdb_id}/" if imdb_id else "https://www.imdb.com/"

        lines.append(
            f'    <a href="{escape(href)}" target="_blank" rel="noopener noreferrer">'
        )
        lines.append(f'    <img src="{escape(img_url)}" />')
        lines.append("    </a>")
        lines.append("")

    lines.append("</div>")
    return "\n".join(lines)


# ── Job entry point ────────────────────────────────────────────────────────────


def run(params: dict, *, config, progress=None) -> dict:
    """Fetch OMDB poster art for the upcoming weekend and generate an MOTD snippet.

    Resolves the workbook via the same SharePoint credentials as the fetchurls job.
    Each movie's poster is saved to ``motd.poster_dir`` as
    ``art-YYYY-MM-DD-movieN-nightM.jpg``. A ``<div class="poster-grid">`` HTML
    snippet is written to ``motd.output_dir/motd-{sheet_name}.html``.
    """
    try:
        import openpyxl as _oxl
    except ImportError:
        raise RuntimeError("openpyxl is required for motd_posters")

    def _emit(detail: dict) -> None:
        if progress:
            progress(detail)

    dry_run = bool(params.get("dry_run", False))

    omdb_key: str = getattr(config, "omdb_api_key", "") or ""
    if not omdb_key:
        raise RuntimeError(
            "omdb_api_key is not configured — motd_posters requires an OMDB API key"
        )

    motd_cfg = getattr(config, "motd", None)
    poster_dir = Path(
        getattr(motd_cfg, "poster_dir", "/home/mediacms.io/mediacms/static/motd_boxes")
    ).expanduser()
    poster_base_url: str = getattr(
        motd_cfg, "poster_base_url", "https://www.dropsugar.co/static/motd_boxes"
    )
    output_dir = Path(getattr(motd_cfg, "output_dir", "~/kryten")).expanduser()
    placeholder_dir: str = getattr(config, "placeholder_dir", "")

    if not dry_run:
        poster_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)

    # Resolve workbook source (mirrors fetchurls.run())
    local_override: str = params.get("workbook_path") or ""
    fu = getattr(config, "fetchurls", None)
    sp_tenant: str = getattr(fu, "sharepoint_tenant_id", "") if fu else ""
    sp_client: str = getattr(fu, "sharepoint_client_id", "") if fu else ""
    sp_share: str = getattr(fu, "sharepoint_sharing_url", "") if fu else ""
    sp_cache: str = getattr(fu, "token_cache_path", "") if fu else ""
    cfg_local: str = getattr(fu, "workbook_path", "") if fu else ""

    use_sharepoint = bool(sp_tenant and sp_client and sp_share) and not local_override

    if use_sharepoint:
        _emit({"phase": "auth"})
        graph_token = acquire_graph_token_silent(sp_tenant, sp_client, sp_cache)
        if not graph_token:
            raise RuntimeError(
                "SharePoint token unavailable — run: "
                "python -m kryten_webqueue.jobs.fetchurls_auth"
            )
        _emit({"phase": "download"})
        wb_bytes, _, _ = download_sharepoint_xlsx(graph_token, sp_share)
    else:
        workbook_path = local_override or cfg_local
        if not workbook_path:
            raise RuntimeError(
                "motd_posters needs a workbook source: configure "
                "fetchurls.sharepoint_* or fetchurls.workbook_path"
            )
        wb_file = Path(workbook_path)
        if not wb_file.exists():
            raise RuntimeError(f"Workbook not found: {wb_file}")
        wb_bytes = wb_file.read_bytes()

    sheet_name, friday, saturday = upcoming_weekend_sheet()
    sunday = friday + datetime.timedelta(days=2)

    wb_peek = _oxl.load_workbook(io.BytesIO(wb_bytes), read_only=True)
    all_sheets = wb_peek.sheetnames
    wb_peek.close()

    if sheet_name not in all_sheets:
        weekend_sheets = [s for s in all_sheets if _SHEET_DATE_RE.match(s.strip())]
        raise RuntimeError(
            f"Weekend sheet '{sheet_name}' not found. "
            f"Available: {', '.join(weekend_sheets or all_sheets)}"
        )

    _emit({"phase": "parsing", "sheet": sheet_name})
    movies_by_section = extract_movies_by_section(wb_bytes, sheet_name)

    _section_date: dict[str, datetime.date] = {
        "friday": friday,
        "saturday-morning": saturday,
        "saturday-night": saturday,
        "sunday-morning": sunday,
        "sunday-daytime": sunday,
    }

    # Build ordered entry list; per-night counters ensure consecutive numbering
    night_counters: dict[int, int] = {}
    night_entries: list[dict] = []
    for slug in _SECTION_ORDER:
        titles = movies_by_section.get(slug, [])
        if not titles:
            continue
        night = _NIGHT_MAP.get(slug, 1)
        date = _section_date.get(slug, friday)
        for title in titles:
            night_counters[night] = night_counters.get(night, 0) + 1
            idx = night_counters[night]
            night_entries.append(
                {
                    "night": night,
                    "date": date,
                    "title": title,
                    "filename": _poster_filename(date, idx, night),
                    "imdb_id": None,
                }
            )

    total = len(night_entries)
    _emit({"phase": "posters", "total": total, "sheet": sheet_name})
    logger.info("motd_posters: %d movie(s) on sheet %s", total, sheet_name)

    resolved = placeholder_used = failed = 0

    for i, entry in enumerate(night_entries):
        title = entry["title"]
        filename = entry["filename"]
        dest = poster_dir / filename

        _emit({"phase": "posters", "movie": title, "index": i + 1, "total": total})

        info = _resolve_omdb(title, api_key=omdb_key)

        if info:
            entry["imdb_id"] = info["imdb_id"]
            if dry_run:
                logger.info(
                    "motd_posters [dry-run]: %s → %s (%s)", title, info["imdb_id"], filename
                )
                resolved += 1
            else:
                img_bytes = _download_image(info["poster_url"])
                if img_bytes:
                    dest.write_bytes(img_bytes)
                    logger.info(
                        "motd_posters: saved %s (%d bytes)", filename, len(img_bytes)
                    )
                    resolved += 1
                else:
                    logger.warning(
                        "motd_posters: poster download failed for %r; using placeholder",
                        title,
                    )
                    ph = _pick_placeholder(placeholder_dir)
                    if ph:
                        dest.write_bytes(ph)
                        placeholder_used += 1
                    else:
                        failed += 1
        else:
            logger.warning("motd_posters: no OMDB result for %r", title)
            if not dry_run:
                ph = _pick_placeholder(placeholder_dir)
                if ph:
                    dest.write_bytes(ph)
                    placeholder_used += 1
                else:
                    failed += 1
            else:
                placeholder_used += 1  # count in dry-run for reporting

    html_snippet = _generate_html(night_entries, poster_base_url)
    output_path = output_dir / f"motd-{sheet_name}.html"

    if dry_run:
        logger.info("motd_posters [dry-run]: HTML snippet (not written):\n%s", html_snippet)
    else:
        output_path.write_text(html_snippet, encoding="utf-8")
        logger.info("motd_posters: wrote HTML snippet → %s", output_path)

    return {
        "sheet": sheet_name,
        "total": total,
        "resolved": resolved,
        "placeholder_used": placeholder_used,
        "failed": failed,
        "output_path": str(output_path),
        "dry_run": dry_run,
    }
