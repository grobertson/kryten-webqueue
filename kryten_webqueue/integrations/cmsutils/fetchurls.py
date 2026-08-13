#!/usr/bin/env python3
# VENDORED from d:\Devel\cmsutils\fetchurls.py on 2026-06-09.
# Adapted for in-process use by kryten-webqueue jobs: a headless
# run(params, *, config, progress) entry point is appended at the bottom. It
# reads a LOCAL workbook (no SharePoint/Graph/MSAL — OQ-1), targets the upcoming
# weekend's sheet, and resolves off-site URLs via the in-process yt-pipe
# downloader instead of shelling out to fetch.ps1. The original CLI main()/
# argparse + SharePoint path is retained but unused by the service.
"""
fetchurls.py - Import off-site media and build dropsugar.co playlist files.

Reads source URLs from an Excel sheet on SharePoint (or a local file / plain
text file), grouped into playlist sections by the section headers found
in column A:

    Friday Schedule           → playlists/{sheet}-friday.txt
    Saturday Schedule         → playlists/{sheet}-saturday-night.txt
    Saturday Morning Cartoons → playlists/{sheet}-saturday-morning.txt
    Sunday Morning            → playlists/{sheet}-sunday-morning.txt
    Sunday Afternoon          → playlists/{sheet}-sunday-daytime.txt

For each source URL:
  • dropsugar.co  — validated with a HEAD request, kept as-is
  • YouTube/Tubi  — processed through ../yt-pipe/fetch.ps1 which downloads
                    the content and uploads it to dropsugar.co

On success the resolved MediaCMS view URLs are written to playlist files that
match the format consumed by the channel injection utility.  Unresolvable URLs
are kept as-is and logged to a separate failures file.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SETUP — SharePoint / Graph API access
──────────────────────────────────────
1. Create an Azure AD App Registration:
     https://portal.azure.com → Azure Active Directory → App registrations → New
     • Name: anything (e.g. "cmsutils-fetchurls")
     • Supported account types: "Accounts in this organizational directory only"
     • Platform: Mobile and desktop applications (tick "https://login.microsoftonline.com/common/oauth2/nativeclient")
     • Leave Redirect URI blank (device code flow does not need one)

2. Under API Permissions → Add:
     Microsoft Graph → Delegated → Files.Read.All   (to read SharePoint files)

3. Under Authentication → tick "Allow public client flows" → Save

4. Copy the Application (client) ID and Directory (tenant) ID from the Overview page.

5. Find your tenant's GUID from the Overview page, *or* set tenant_id to
   "curiousmotors.onmicrosoft.com" — both work.

6. Add to config.yaml:

     sharepoint:
       tenant_id:   "YOUR-TENANT-ID-OR-DOMAIN"
       client_id:   "YOUR-APP-CLIENT-ID"
       sharing_url: "https://curiousmotors.sharepoint.com/:x:/r/..."

     fetch:
       script: "../yt-pipe/fetch.ps1"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Usage
─────
  # Pull from SharePoint and process the sheet for this weekend (auto-detected)
  python fetchurls.py

  # Process a specific sheet by name
  python fetchurls.py --sheet "3.6-3.7"

  # Use a locally synced / downloaded Excel file
  python fetchurls.py --file "C:/path/to/Channel Z Playlist.xlsx" --sheet "3.6-3.7"

  # Just process a text file of raw URLs (one per line, no section splitting)
  python fetchurls.py --text urls.txt --label "3.6-3.7"

  # Dry-run: show what would happen without calling fetch.ps1
  python fetchurls.py --sheet "3.6-3.7" --dry-run

Output
──────
  playlists/{sheet}-friday.txt
  playlists/{sheet}-saturday-night.txt
  playlists/{sheet}-saturday-morning.txt
  playlists/{sheet}-sunday-morning.txt
  playlists/{sheet}-sunday-daytime.txt
  playlists/{sheet}-failures.txt          (if any URLs could not be resolved)
"""

from __future__ import annotations

import argparse
import base64
import datetime
import io
import json
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import logging
import requests
import yaml

logger = logging.getLogger(__name__)

# ── Optional dependencies (checked at runtime) ────────────────────────────────
try:
    import openpyxl

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import msal

    _HAS_MSAL = True
except ImportError:
    _HAS_MSAL = False


# ── Constants ─────────────────────────────────────────────────────────────────

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
# Files.ReadWrite.All is needed for writing resolved URLs back to column F.
# On first run after upgrading from Files.Read.All the user will be asked
# to re-consent once interactively; subsequent runs use the cached token.
GRAPH_SCOPES = ["https://graph.microsoft.com/Files.ReadWrite.All"]

# Column indices in the Excel sheet (0-based when using openpyxl cell.column)
COL_SECTION = 1  # Column A — section headers
COL_URL = 5  # Column E — source URLs

# Section-name substrings → output playlist slug.
# Order matters: more specific keys must precede broader ones (see
# _classify_section, which returns the first substring match).
SECTION_MAP = {
    "friday": "friday",
    "saturday morning": "saturday-morning",
    "saturday": "saturday-night",  # must come AFTER "saturday morning"
    "sunday morning": "sunday-morning",
    "sunday afternoon": "sunday-daytime",
}

DROPSUGAR_HOST = "dropsugar.co"
OFFSITE_HOSTS = ("youtube.com", "youtu.be", "tubi.tv", "tubitv.com")

REQUEST_TIMEOUT = 20  # seconds for HEAD validation
FETCH_TIMEOUT = 900  # 15 min per URL (download + encode + upload)

# ── Data structures ───────────────────────────────────────────────────────────


@dataclass
class Config:
    fetch_script: str = "../yt-pipe/fetch.ps1"
    api_url: str = "https://www.dropsugar.co/api/v1"
    api_token: str = ""
    sp_tenant_id: str = ""
    sp_client_id: str = ""
    sp_sharing_url: str = ""
    sp_token_cache: str = ".fetchurls_tokens.bin"  # persisted MSAL cache


@dataclass
class ProcessResult:
    original_url: str
    resolved_url: str
    success: bool
    note: str = ""
    row_number: int = 0  # 1-based Excel row; 0 = not from a spreadsheet


@dataclass
class WritebackContext:
    """Holds Graph API credentials needed to write column F incrementally."""

    token: str
    drive_id: str
    item_id: str
    sheet_name: str
    dry_run: bool = False
    session_id: str = ""  # Graph workbook session for reliable writes
    # Auth details for automatic token refresh (tokens expire after ~1 hour)
    tenant_id: str = ""
    client_id: str = ""
    cache_path: str = ""
    # Tracking
    writes_ok: int = 0
    writes_fail: int = 0


# ── Configuration ─────────────────────────────────────────────────────────────


def load_config(config_path: Path) -> Config:
    cfg = Config()
    if not config_path.exists():
        return cfg
    with config_path.open() as f:
        raw = yaml.safe_load(f) or {}

    api = raw.get("api", {})
    cfg.api_url = api.get("url", cfg.api_url)
    cfg.api_token = api.get("token", cfg.api_token)

    sp = raw.get("sharepoint", {})
    cfg.sp_tenant_id = sp.get("tenant_id", cfg.sp_tenant_id)
    cfg.sp_client_id = sp.get("client_id", cfg.sp_client_id)
    cfg.sp_sharing_url = sp.get("sharing_url", cfg.sp_sharing_url)
    cfg.sp_token_cache = sp.get("token_cache", cfg.sp_token_cache)

    fetch = raw.get("fetch", {})
    cfg.fetch_script = fetch.get("script", cfg.fetch_script)

    return cfg


# ── SharePoint / Graph API ────────────────────────────────────────────────────


def _check_msal():
    if not _HAS_MSAL:
        sys.exit("ERROR: 'msal' package not installed.\n" "  pip install msal\n")


def _check_openpyxl():
    if not _HAS_OPENPYXL:
        sys.exit(
            "ERROR: 'openpyxl' package not installed.\n" "  pip install openpyxl\n"
        )


def acquire_graph_token(tenant_id: str, client_id: str, cache_path: str = "") -> str:
    """
    Authenticate via MSAL and return a Graph access token.

    Uses a persistent SerializableTokenCache stored on disk so that
    subsequent runs (including cron jobs) authenticate silently via the
    cached refresh token.  The refresh token is valid for 90 days by
    default (up to 1 year with Entra ID Continuous Access Evaluation).
    On first run, or after token expiry, a device code prompt is shown.

    token_cache_path — path to the binary cache file; defaults to
    .fetchurls_tokens.bin in the same directory as this script.
    """
    _check_msal()

    # Resolve cache file path
    if not cache_path:
        cache_path = str(Path(__file__).parent / ".fetchurls_tokens.bin")
    cache_file = Path(cache_path)

    # Load existing cache from disk
    token_cache = msal.SerializableTokenCache()
    if cache_file.exists():
        token_cache.deserialize(cache_file.read_text(encoding="utf-8"))

    app = msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        token_cache=token_cache,
    )

    def _save_cache():
        if token_cache.has_state_changed:
            cache_file.write_text(token_cache.serialize(), encoding="utf-8")

    # Try a silent token first (uses the on-disk cache / refresh token)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])
        if result and "access_token" in result:
            _save_cache()
            return result["access_token"]

    # No valid cached token — fall back to device code flow
    flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
    if "user_code" not in flow:
        sys.exit(f"ERROR: Failed to create device flow: {flow}")

    print()
    print("━" * 60)
    print(flow["message"])  # "Go to https://... and enter code ABCDEFGH"
    print("━" * 60)
    print()

    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        sys.exit(
            f"ERROR: Authentication failed: {result.get('error_description', result)}"
        )
    _save_cache()
    return result["access_token"]


def acquire_graph_token_silent(
    tenant_id: str, client_id: str, cache_path: str
) -> Optional[str]:
    """Return a Graph access token from a pre-seeded MSAL cache, or None.

    Unlike :func:`acquire_graph_token`, this NEVER prompts interactively — it is
    safe to call from the headless webqueue service. The cache must have been
    seeded out-of-band (``python -m kryten_webqueue.jobs.fetchurls_auth``). A
    None return means the admin must (re)authenticate.
    """
    if not _HAS_MSAL:
        raise RuntimeError(
            "fetchurls SharePoint integration requires 'msal'. "
            "Install the optional extra: pip install 'kryten-webqueue[jobs]'"
        )
    if not cache_path:
        return None
    cache_file = Path(cache_path)
    if not cache_file.exists():
        return None
    token_cache = msal.SerializableTokenCache()
    token_cache.deserialize(cache_file.read_text(encoding="utf-8"))
    app = msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        token_cache=token_cache,
    )
    accounts = app.get_accounts()
    if not accounts:
        return None
    result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])
    if token_cache.has_state_changed:
        cache_file.write_text(token_cache.serialize(), encoding="utf-8")
    if result and "access_token" in result:
        return result["access_token"]
    return None


def _encode_sharing_url(url: str) -> str:
    """Encode a SharePoint sharing/document URL for use with the Graph shares API."""
    encoded = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")
    return f"u!{encoded}"


def download_sharepoint_xlsx(token: str, sharing_url: str) -> tuple[bytes, str, str]:
    """
    Download the Excel file from SharePoint via the Graph shares endpoint.

    Returns: (file_bytes, drive_id, item_id)
    drive_id and item_id are needed to write back resolved URLs via the
    Graph Excel workbook API.
    """
    encoded = _encode_sharing_url(sharing_url)
    headers = {"Authorization": f"Bearer {token}"}

    # Resolve the sharing URL to a driveItem
    meta_url = f"{GRAPH_BASE}/shares/{encoded}/driveItem"
    r = requests.get(meta_url, headers=headers, timeout=REQUEST_TIMEOUT)
    if r.status_code != 200:
        raise RuntimeError(
            f"Could not resolve SharePoint file via Graph API (HTTP {r.status_code}). "
            f"Response: {r.text[:500]}"
        )
    item = r.json()

    # Always capture drive/item IDs for later writeback
    drive_id = item.get("parentReference", {}).get("driveId") or item.get(
        "remoteItem", {}
    ).get("parentReference", {}).get("driveId", "")
    item_id = item.get("id", "")

    download_url = item.get("@microsoft.graph.downloadUrl") or item.get("downloadUrl")

    if not download_url:
        # Fetch download URL by following the content redirect
        dl_r = requests.get(
            f"{GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content",
            headers=headers,
            allow_redirects=False,
            timeout=REQUEST_TIMEOUT,
        )
        if dl_r.status_code == 302:
            download_url = dl_r.headers["Location"]
        else:
            raise RuntimeError(
                f"Could not get SharePoint download URL (HTTP {dl_r.status_code}). "
                f"Response: {dl_r.text[:500]}"
            )

    content_r = requests.get(download_url, timeout=60)
    if not content_r.ok:
        raise RuntimeError(
            f"SharePoint file download failed (HTTP {content_r.status_code}). "
            f"Response: {content_r.text[:500]}"
        )
    return content_r.content, drive_id, item_id


# ── Sheet auto-detection ─────────────────────────────────────────────────────

# Matches sheet names like "3.6-3.7", "3.13-3.14", "2.7 - 2.8"
_SHEET_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\s*-\s*\d{1,2}\.\d{1,2}$")


def _auto_select_sheet(sheet_names: list[str]) -> Optional[str]:
    """
    Pick the sheet whose start date (Friday) is the most recent one on or
    before today — i.e. the current or just-past broadcast weekend.

    Falls back to the nearest upcoming date sheet, then the raw last sheet.
    """
    today = datetime.date.today()
    year = today.year

    candidates: list[tuple[datetime.date, str]] = []
    for name in sheet_names:
        m = _SHEET_DATE_RE.match(name.strip())
        if not m:
            continue
        month, day = int(m.group(1)), int(m.group(2))
        try:
            d = datetime.date(year, month, day)
        except ValueError:
            continue
        candidates.append((d, name))

    if not candidates:
        # No date-formatted sheets — fall back to last sheet in workbook
        return None

    candidates.sort(key=lambda x: x[0])

    # Most recent start date on or before today (current/past weekend)
    past = [(d, n) for d, n in candidates if d <= today]
    if past:
        return past[-1][1]

    # All sheets are future dates — return the nearest upcoming one
    return candidates[0][1]


# ── Excel parsing ─────────────────────────────────────────────────────────────


def _classify_section(cell_value: str) -> Optional[str]:
    """
    Return the slug for a section header cell, or None if it's not a header.
    Checks in order so 'saturday morning' matches before plain 'saturday'.
    """
    if not cell_value:
        return None
    v = cell_value.strip().lower()
    for keyword, slug in SECTION_MAP.items():
        if keyword in v:
            return slug
    return None


def parse_excel_sections(
    wb_bytes: bytes,
    sheet_name: str,
) -> dict[str, list[tuple[int, str, str]]]:
    """
    Parse the workbook and return (row_number, url, col_f) triples grouped
    by section slug.

    Scans column A for section headers, column E for source URLs, and
    column F for previously-resolved URLs (from an earlier run).
    Row numbers are 1-based (matching Excel row numbers) for use with the
    Graph Excel API when writing resolved URLs back to column F.

    Returns dict: {
        "friday":           [(row, url, col_f), ...],
        "saturday-night":   [(row, url, col_f), ...],
        "saturday-morning": [(row, url, col_f), ...],
        "sunday-morning":   [(row, url, col_f), ...],
        "sunday-daytime":   [(row, url, col_f), ...],
    }
    """
    _check_openpyxl()
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        sys.exit(
            f"ERROR: Sheet '{sheet_name}' not found.\n" f"Available sheets: {available}"
        )

    ws = wb[sheet_name]
    sections: dict[str, list[tuple[int, str, str]]] = {
        "friday": [],
        "saturday-night": [],
        "saturday-morning": [],
        "sunday-morning": [],
        "sunday-daytime": [],
    }
    current_section: Optional[str] = None

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Column A is index 0, Column E is index 4, Column F is index 5
        col_a = str(row[0]).strip() if row[0] is not None else ""
        col_e = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        col_f = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

        slug = _classify_section(col_a)
        if slug:
            current_section = slug
            continue

        if current_section and col_e and col_e.lower().startswith("http"):
            sections[current_section].append((row_num, col_e, col_f))

    wb.close()
    return sections


def parse_text_file(path: Path) -> list[str]:
    """Read one URL per line from a text file, skipping blanks and # comments."""
    urls = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#"):
            urls.append(line)
    return urls


# ── URL classification ────────────────────────────────────────────────────────


def is_dropsugar(url: str) -> bool:
    return DROPSUGAR_HOST in url.lower()


def is_offsite(url: str) -> bool:
    lower = url.lower()
    return any(h in lower for h in OFFSITE_HOSTS)


def validate_dropsugar(url: str) -> tuple[bool, str]:
    """Validate a dropsugar.co URL.  Returns (ok, note).

    Manifest API URLs (/api/v1/media/cytube/{key}.json) return HTTP 500 on
    HEAD requests due to a server quirk, but GET works fine and the JSON
    payload is small.  All other URLs are checked with HEAD.
    """
    try:
        if "/api/v1/media/" in url:
            r = requests.get(url, timeout=REQUEST_TIMEOUT)
        else:
            r = requests.head(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code < 400:
            return True, f"HTTP {r.status_code}"
        return False, f"HTTP {r.status_code}"
    except requests.RequestException as exc:
        return False, str(exc)


# ── fetch.ps1 invocation ──────────────────────────────────────────────────────


def _is_wsl() -> bool:
    """Return True when running inside Windows Subsystem for Linux."""
    try:
        return "microsoft" in Path("/proc/version").read_text(encoding="utf-8").lower()
    except OSError:
        return False


def _to_windows_path(p: Path) -> str:
    """
    Convert a Linux path to a Windows path string when running under WSL.

    Uses `wslpath -w` so that Windows processes (powershell.exe) can open
    the file.  Falls back to the string representation of the path if
    wslpath is unavailable.
    """
    try:
        result = subprocess.run(
            ["wslpath", "-w", str(p)],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return str(p)


def _resolve_fetch_script(script_path: str) -> Path:
    """Resolve the fetch script path relative to this file's directory."""
    p = Path(script_path)
    if not p.is_absolute():
        p = (Path(__file__).parent / p).resolve()
    return p


def _strip_playlist_params(url: str) -> str:
    """Remove &list= and &index= params from YouTube URLs.

    When a YouTube URL contains a playlist reference the downstream
    pipeline may try to process every video in the list.  We only want
    the single video identified by the ?v= parameter.
    """
    if "list=" not in url:
        return url
    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

    p = urlparse(url)
    qs = parse_qs(p.query, keep_blank_values=True)
    qs.pop("list", None)
    qs.pop("index", None)
    cleaned = urlunparse(p._replace(query=urlencode(qs, doseq=True)))
    if cleaned != url:
        print("    (stripped playlist params from URL)")
    return cleaned


def run_fetch(
    url: str, fetch_script: Path, dry_run: bool = False
) -> tuple[bool, str, str]:
    """
    Run ../yt-pipe/fetch.ps1 with the given URL.

    Returns:
        (success, resolved_url, note)
        resolved_url is the MediaCMS view URL on success, empty string on failure.
    """
    if dry_run:
        print(f"    [DRY-RUN] Would run: fetch.ps1 {url}")
        return True, "https://www.dropsugar.co/view?m=DRY_RUN", "dry-run"

    # Strip YouTube playlist params so fetch.ps1 processes only the single video
    url = _strip_playlist_params(url)

    # Under WSL, powershell.exe is a Windows process and cannot open /mnt/...
    # paths — convert to a Windows path (e.g. D:\Devel\yt-pipe\fetch.ps1).
    script_arg = _to_windows_path(fetch_script) if _is_wsl() else str(fetch_script)

    cmd = [
        "powershell.exe",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        script_arg,
        url,
    ]

    print("    Running fetch.ps1 …", flush=True)
    # Force UTF-8 mode for any Python child processes spawned by fetch.ps1.
    # Without this, piped stdout on Windows defaults to cp1252 which chokes
    # on emoji / non-Latin characters in youtube_to_mediacms.py output.
    env = {**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"}
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=FETCH_TIMEOUT if FETCH_TIMEOUT > 0 else None,
            cwd=fetch_script.parent,  # fetch.ps1 uses relative paths internally
            env=env,
        )
    except subprocess.TimeoutExpired:
        return False, "", "fetch.ps1 timed out"
    except FileNotFoundError:
        return False, "", "powershell.exe not found"

    stdout = result.stdout
    stderr = result.stderr

    # Parse the success block — two possible formats:
    #
    # 1. New upload:
    #   MediaCMS URL: https://www.dropsugar.co/view?m=XXXXXXXX
    #   Manifest URL: https://www.dropsugar.co/api/v1/media/cytube/XXXXXXXX.json?format=json
    #
    # 2. Already exists (metadata enriched):
    #   MediaCMS ID: XXXXXXXX
    media_url = _extract_line(stdout, r"MediaCMS URL:\s*(https://\S+)")
    manifest_url = _extract_line(stdout, r"Manifest URL:\s*(https://\S+)")

    if media_url:
        return True, media_url, f"manifest: {manifest_url or '(not found)'}"

    # Check for "already exists" output — has MediaCMS ID but no full URL
    media_id = _extract_line(stdout, r"MediaCMS ID:\s*(\S+)")
    if media_id:
        view_url = f"https://www.dropsugar.co/view?m={media_id}"
        return True, view_url, "already exists (metadata enriched)"

    # Identify common failure reasons for better log messages
    note = _classify_failure(stdout + stderr)
    if result.returncode != 0 and not note:
        note = f"exit code {result.returncode}"

    # Emit tail of stdout so the operator can diagnose inline
    tail = "\n".join((stdout + stderr).strip().splitlines()[-8:])
    if tail:
        print(f"    --- fetch output (tail) ---\n{tail}\n    ---", flush=True)

    return False, "", note or "unknown failure"


def _extract_line(text: str, pattern: str) -> Optional[str]:
    m = re.search(pattern, text)
    return m.group(1).strip() if m else None


def _classify_failure(text: str) -> str:
    lower = text.lower()
    if (
        "sign in to confirm" in lower
        or "age-restricted" in lower
        or "age restricted" in lower
    ):
        return "age-restricted (YouTube) — cookies required"
    if "drm" in lower or "widevine" in lower or "not available" in lower:
        return "DRM-protected or geo-restricted"
    if "video unavailable" in lower:
        return "video unavailable"
    if "private video" in lower:
        return "private video"
    if "already imported" in lower:
        # Script found it in MediaCMS already — treat as success if URL present
        return ""
    return ""


# ── Processing ────────────────────────────────────────────────────────────────


def process_url(
    url: str,
    fetch_script: Path,
    dry_run: bool = False,
    validate: bool = True,
    row_number: int = 0,
) -> ProcessResult:
    """Resolve a single URL to a dropsugar.co view URL."""

    if is_dropsugar(url):
        if validate:
            ok, note = validate_dropsugar(url)
            if ok:
                return ProcessResult(
                    url, url, True, f"dropsugar (validated: {note})", row_number
                )
            else:
                return ProcessResult(
                    url, url, False, f"dropsugar validation failed: {note}", row_number
                )
        return ProcessResult(url, url, True, "dropsugar (not validated)", row_number)

    if is_offsite(url):
        ok, resolved, note = run_fetch(url, fetch_script, dry_run=dry_run)
        if ok:
            return ProcessResult(url, resolved, True, note, row_number)
        else:
            return ProcessResult(url, url, False, note, row_number)

    # Unknown host — keep as-is but warn
    return ProcessResult(url, url, True, "unrecognised host — kept as-is", row_number)


def process_section(
    name: str,
    url_rows: list[tuple[int, str, str]],
    fetch_script: Path,
    dry_run: bool,
    validate: bool,
    wb_ctx: Optional[WritebackContext] = None,
) -> list[ProcessResult]:
    results = []
    total = len(url_rows)
    revalidated = 0
    print(f"\n  ── {name.upper()} ({total} URLs) ──")
    for i, (row_num, url, col_f) in enumerate(url_rows, 1):
        short = url if len(url) <= 70 else url[:67] + "…"
        print(f"  [{i:>2}/{total}] {short}")

        # If column F already has a resolved URL from a previous run:
        #   • validate=True  → re-check it; fall through to re-process if stale
        #   • validate=False → trust it as-is (--no-validate speeds up reruns)
        if (
            col_f
            and DROPSUGAR_HOST in col_f.lower()
            and col_f.lower().startswith("http")
        ):
            if not validate:
                r = ProcessResult(
                    url, col_f, True, "col F kept (validation skipped)", row_num
                )
                print(f"         ✓  {r.note}")
                results.append(r)
                revalidated += 1
                continue
            ok, note = validate_dropsugar(col_f)
            if ok:
                r = ProcessResult(
                    url, col_f, True, f"col F revalidated ({note})", row_num
                )
                print(f"         ✓  {r.note}")
                results.append(r)
                revalidated += 1
                continue
            # Stale — fall through and re-process col_e
            print(f"         ⚠  col F no longer valid ({note}), re-processing…")

        r = process_url(
            url, fetch_script, dry_run=dry_run, validate=validate, row_number=row_num
        )
        status = "✓" if r.success else "✗"
        print(f"         {status}  {r.note}")
        results.append(r)

        # Write resolved URL to column F immediately so progress survives interrupts
        if r.success and r.row_number > 0 and wb_ctx is not None:
            _write_cell_f(wb_ctx, r)

    if revalidated:
        print(f"  ({revalidated} already resolved, revalidated from column F)")
    return results


# ── SharePoint writeback ─────────────────────────────────────────────────────

WRITEBACK_MAX_RETRIES = 4  # retry transient Graph API failures
WRITEBACK_DELAY = 1.0  # seconds between writes (avoids rate limits)
WRITEBACK_TIMEOUT = 30  # per-request timeout (Excel API can be slow)


def _workbook_api_base(ctx: WritebackContext) -> str:
    """Return the common Graph API prefix for this workbook."""
    return f"{GRAPH_BASE}/drives/{ctx.drive_id}/items/{ctx.item_id}/workbook"


def _refresh_token(ctx: WritebackContext) -> bool:
    """
    Re-acquire the Graph API access token using the cached refresh token.

    Access tokens expire after ~1 hour.  Long runs that process many URLs
    through fetch.ps1 (up to 15 min each) will outlast a single token.
    This silently refreshes via the on-disk MSAL cache.

    Returns True if the token was refreshed, False on failure.
    """
    if not ctx.tenant_id or not ctx.client_id:
        return False
    try:
        new_token = acquire_graph_token(ctx.tenant_id, ctx.client_id, ctx.cache_path)
        if new_token and new_token != ctx.token:
            ctx.token = new_token
            print("    (refreshed Graph API token)")
            return True
    except SystemExit:
        # acquire_graph_token calls sys.exit on failure — catch it
        pass
    return False


def _open_fresh_session(ctx: WritebackContext) -> None:
    """
    Close any existing session and open a new one with persistChanges=True.

    The Graph Excel API silently discards writes that arrive on an expired
    session (returns HTTP 200 but does not persist).  Sessions time out
    after ~5 min of inactivity, and fetch.ps1 calls between writes can
    easily exceed that.  Opening a fresh session right before each write
    guarantees the session is live.

    Closing the old session first avoids hitting the ~5 concurrent session
    limit per workbook.
    """
    # Close old session (best-effort, ignore errors)
    if ctx.session_id:
        close_url = f"{_workbook_api_base(ctx)}/closeSession"
        try:
            requests.post(
                close_url,
                headers={
                    "Authorization": f"Bearer {ctx.token}",
                    "Content-Type": "application/json",
                    "workbook-session-id": ctx.session_id,
                },
                data="{}",
                timeout=WRITEBACK_TIMEOUT,
            )
        except requests.RequestException:
            pass
        ctx.session_id = ""

    # Create new session
    url = f"{_workbook_api_base(ctx)}/createSession"
    headers = {
        "Authorization": f"Bearer {ctx.token}",
        "Content-Type": "application/json",
    }
    body = json.dumps({"persistChanges": True})
    try:
        resp = requests.post(url, headers=headers, data=body, timeout=WRITEBACK_TIMEOUT)
        if resp.status_code in (200, 201):
            sid = resp.json().get("id", "")
            if sid:
                ctx.session_id = sid
                return
        # 401 means token expired — try refresh once
        if resp.status_code == 401 and _refresh_token(ctx):
            headers["Authorization"] = f"Bearer {ctx.token}"
            resp = requests.post(
                url, headers=headers, data=body, timeout=WRITEBACK_TIMEOUT
            )
            if resp.status_code in (200, 201):
                sid = resp.json().get("id", "")
                if sid:
                    ctx.session_id = sid
                    return
    except requests.RequestException:
        pass
    # Fall back to no session — writes will still be attempted but may not persist
    ctx.session_id = ""


def _close_workbook_session(ctx: WritebackContext) -> None:
    """Close a previously opened workbook session (best-effort)."""
    if not ctx.session_id:
        return
    url = f"{_workbook_api_base(ctx)}/closeSession"
    headers = {
        "Authorization": f"Bearer {ctx.token}",
        "Content-Type": "application/json",
        "workbook-session-id": ctx.session_id,
    }
    try:
        resp = requests.post(url, headers=headers, data="{}", timeout=WRITEBACK_TIMEOUT)
        if resp.status_code in (200, 204):
            print("  Closed workbook session")
        else:
            print(f"  ⚠ closeSession HTTP {resp.status_code}: {resp.text[:120]}")
    except requests.RequestException as exc:
        print(f"  ⚠ closeSession error: {exc}")
    ctx.session_id = ""


def _write_cell_f(ctx: WritebackContext, r: ProcessResult) -> bool:
    """
    Write a single resolved URL to column F in the SharePoint workbook.

    Opens a fresh session (with persistChanges=True) immediately before
    each write to guarantee it is live — the Graph API silently discards
    writes on expired sessions (HTTP 200 but no persistence).  On 401
    (token expired) the access token is refreshed via the MSAL cache.
    Retries on 409/423/429/503 with exponential backoff.

    Returns True on success, False on error (errors are logged but do not
    stop processing).
    """
    if ctx.dry_run:
        print(f"    [DRY-RUN] writeback F{r.row_number} = {r.resolved_url}")
        return True

    if not ctx.drive_id or not ctx.item_id:
        print(f"    ⚠ writeback F{r.row_number}: no drive/item ID — skipped")
        return False

    # Open a fresh session right before the write so it can't be stale
    _open_fresh_session(ctx)

    sheet_enc = ctx.sheet_name.replace("'", "''")
    address = f"F{r.row_number}"
    url = (
        f"{_workbook_api_base(ctx)}"
        f"/worksheets('{sheet_enc}')/range(address='{address}')"
    )
    body = json.dumps({"values": [[r.resolved_url]]})

    for attempt in range(1, WRITEBACK_MAX_RETRIES + 1):
        headers = {
            "Authorization": f"Bearer {ctx.token}",
            "Content-Type": "application/json",
        }
        if ctx.session_id:
            headers["workbook-session-id"] = ctx.session_id

        try:
            resp = requests.patch(
                url, headers=headers, data=body, timeout=WRITEBACK_TIMEOUT
            )

            if resp.status_code in (200, 204):
                short_url = (
                    r.resolved_url
                    if len(r.resolved_url) <= 60
                    else r.resolved_url[:57] + "…"
                )
                print(f"    ✓ wrote F{r.row_number} = {short_url}")
                ctx.writes_ok += 1
                time.sleep(WRITEBACK_DELAY)
                return True

            # Token expired — refresh and retry immediately
            if resp.status_code == 401:
                print(f"    ⚠ writeback F{r.row_number}: HTTP 401 (token expired)")
                if _refresh_token(ctx):
                    _open_fresh_session(ctx)
                    continue  # retry with new token + session
                print(f"    ✗ writeback F{r.row_number}: could not refresh token")
                break

            # Session expired or invalid
            if resp.status_code == 404:
                print(f"    ⚠ writeback F{r.row_number}: HTTP 404 (session expired?)")
                _open_fresh_session(ctx)
                continue  # retry with new session

            # Retryable server-side errors
            if (
                resp.status_code in (409, 423, 429, 503)
                and attempt < WRITEBACK_MAX_RETRIES
            ):
                wait = 2**attempt
                retry_after = resp.headers.get("Retry-After")
                if retry_after:
                    try:
                        wait = max(wait, int(retry_after))
                    except ValueError:
                        pass
                print(
                    f"    ⚠ writeback F{r.row_number}: HTTP {resp.status_code}, "
                    f"retrying in {wait}s ({attempt}/{WRITEBACK_MAX_RETRIES})"
                )
                time.sleep(wait)
                continue

            # Non-retryable error
            print(
                f"    ✗ writeback F{r.row_number}: HTTP {resp.status_code} — "
                f"{resp.text[:200]}"
            )
            break

        except requests.RequestException as exc:
            if attempt < WRITEBACK_MAX_RETRIES:
                wait = 2**attempt
                print(
                    f"    ⚠ writeback F{r.row_number}: {exc}, "
                    f"retrying in {wait}s ({attempt}/{WRITEBACK_MAX_RETRIES})"
                )
                time.sleep(wait)
                continue
            print(f"    ✗ writeback F{r.row_number}: {exc}")
            break

    ctx.writes_fail += 1
    return False


# ── Output ────────────────────────────────────────────────────────────────────

SECTION_SLUGS = {
    "friday": "friday",
    "saturday-night": "saturday-night",
    "saturday-morning": "saturday-morning",
    "sunday-morning": "sunday-morning",
    "sunday-daytime": "sunday-daytime",
}

SECTION_LABELS = {
    "friday": "Friday Night",
    "saturday-night": "Saturday Night",
    "saturday-morning": "Saturday Morning",
    "sunday-morning": "Sunday Morning",
    "sunday-daytime": "Sunday Daytime",
}


def write_playlist(path: Path, results: list[ProcessResult]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in results:
            f.write(r.resolved_url + "\n")
    print(f"  Wrote {len(results)} URLs → {path}")


def write_failures(path: Path, all_results: dict[str, list[ProcessResult]]) -> int:
    failures = [
        (section, r)
        for section, results in all_results.items()
        for r in results
        if not r.success
    ]
    if not failures:
        return 0

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        f.write("# fetchurls failure log\n")
        f.write(f"# {len(failures)} unresolved URL(s)\n\n")
        for section, r in failures:
            f.write(f"# [{SECTION_LABELS.get(section, section)}] {r.note}\n")
            f.write(f"{r.original_url}\n\n")
    print(f"  Wrote {len(failures)} failure(s) → {path}")
    return len(failures)


# ── Main ──────────────────────────────────────────────────────────────────────


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Import off-site media and build dropsugar.co playlist files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    src = p.add_mutually_exclusive_group()
    src.add_argument(
        "--file",
        metavar="PATH",
        help="Path to a local Excel workbook (instead of downloading from SharePoint)",
    )
    src.add_argument(
        "--text",
        metavar="PATH",
        help="Path to a plain-text file with one URL per line (skips Excel parsing)",
    )

    p.add_argument(
        "--sheet",
        metavar="NAME",
        help='Sheet name in the workbook, e.g. "3.6-3.7". '
        "Defaults to the most recent sheet.",
    )
    p.add_argument(
        "--label",
        metavar="NAME",
        help="Label to use for output filenames when --text is used (e.g. '3.6-3.7').",
    )
    p.add_argument(
        "--config",
        metavar="PATH",
        default="config.yaml",
        help="Path to config file (default: config.yaml)",
    )
    p.add_argument(
        "--out-dir",
        metavar="DIR",
        default="playlists",
        help="Directory for output playlist files (default: playlists/)",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would happen without running fetch.ps1",
    )
    p.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip HEAD validation of existing dropsugar.co URLs",
    )
    p.add_argument(
        "--section",
        metavar="SLUG",
        choices=list(SECTION_SLUGS),
        help="Process only one section (friday / saturday-night / saturday-morning)",
    )
    p.add_argument(
        "--no-writeback",
        action="store_true",
        help="Skip writing resolved URLs back to column F in the SharePoint spreadsheet",
    )
    return p


def main() -> None:
    args = build_arg_parser().parse_args()

    # ── Config ────────────────────────────────────────────────────────────────
    config_path = Path(args.config)
    cfg = load_config(config_path)

    print()
    print("  ╔══════════════════════════════════════════╗")
    print("  ║   fetchurls — Playlist URL Importer      ║")
    print("  ╚══════════════════════════════════════════╝")
    print()

    if args.dry_run:
        print("  ⚠  DRY-RUN mode — fetch.ps1 will NOT be called")
        print()

    # ── Resolve fetch script ──────────────────────────────────────────────────
    fetch_script = _resolve_fetch_script(cfg.fetch_script)
    if not args.dry_run and not fetch_script.exists():
        sys.exit(
            f"ERROR: fetch script not found: {fetch_script}\n"
            f"  Set fetch.script in config.yaml or check the path."
        )
    print(f"  Fetch script : {fetch_script}")

    # ── Load URLs ─────────────────────────────────────────────────────────────

    if args.text:
        # ── Text-file mode ────────────────────────────────────────────────────
        text_path = Path(args.text)
        if not text_path.exists():
            sys.exit(f"ERROR: Text file not found: {text_path}")

        label = args.label or text_path.stem
        urls = parse_text_file(text_path)
        print(f"  Source       : {text_path} ({len(urls)} URLs)")
        print(f"  Label        : {label}")
        print()

        url_rows = [(0, u, "") for u in urls]  # no Excel row numbers in text-file mode
        results = process_section(
            "all", url_rows, fetch_script, args.dry_run, not args.no_validate
        )

        out_dir = Path(args.out_dir)
        write_playlist(out_dir / f"{label}-imported.txt", results)
        write_failures(out_dir / f"{label}-failures.txt", {"all": results})

    else:
        # ── Excel mode (SharePoint or local file) ─────────────────────────────
        _check_openpyxl()

        sp_drive_id = ""
        sp_item_id = ""
        token = ""

        if args.file:
            file_path = Path(args.file)
            if not file_path.exists():
                sys.exit(f"ERROR: Excel file not found: {file_path}")
            wb_bytes = file_path.read_bytes()
            print(f"  Source       : {file_path}")
        else:
            # Download from SharePoint
            if not cfg.sp_tenant_id or not cfg.sp_client_id:
                sys.exit(
                    "ERROR: sharepoint.tenant_id and sharepoint.client_id must be set "
                    "in config.yaml to download from SharePoint.\n"
                    "  See the setup instructions at the top of fetchurls.py.\n"
                    "  Or use --file to supply a local Excel file."
                )
            if not cfg.sp_sharing_url:
                sys.exit("ERROR: sharepoint.sharing_url must be set in config.yaml.")

            print("  Source       : SharePoint (Graph API)")
            print("  Authenticating …")
            token = acquire_graph_token(
                cfg.sp_tenant_id, cfg.sp_client_id, cfg.sp_token_cache
            )
            print("  Downloading workbook …")
            wb_bytes, sp_drive_id, sp_item_id = download_sharepoint_xlsx(
                token, cfg.sp_sharing_url
            )
            print(f"  Downloaded   : {len(wb_bytes):,} bytes")

        # Determine sheet name
        import openpyxl as _oxl

        wb_peek = _oxl.load_workbook(io.BytesIO(wb_bytes), read_only=True)
        all_sheets = wb_peek.sheetnames
        wb_peek.close()

        sheet_name = args.sheet
        if not sheet_name:
            auto = _auto_select_sheet(all_sheets)
            if auto:
                sheet_name = auto
                print(
                    f"  Sheet        : '{sheet_name}' (auto-selected, current weekend)"
                )
            else:
                sheet_name = all_sheets[-1]
                print(
                    f"  Sheet        : '{sheet_name}' (auto-selected, last in workbook)"
                )
        else:
            print(f"  Sheet        : '{sheet_name}'")

        print(f"  Available    : {', '.join(all_sheets)}")
        print()

        # Parse sections
        sections = parse_excel_sections(wb_bytes, sheet_name)
        total_urls = sum(len(v) for v in sections.values())
        print(
            f"  Found {total_urls} URLs across {sum(1 for v in sections.values() if v)} section(s)"
        )

        # Filter to one section if requested
        if args.section:
            sections = {k: v for k, v in sections.items() if k == args.section}

        # Build writeback context (None if local file or writeback disabled)
        wb_ctx: Optional[WritebackContext] = None
        if not args.no_writeback and not args.file and sp_drive_id and sp_item_id:
            wb_ctx = WritebackContext(
                token=token,
                drive_id=sp_drive_id,
                item_id=sp_item_id,
                sheet_name=sheet_name,
                dry_run=args.dry_run,
                tenant_id=cfg.sp_tenant_id,
                client_id=cfg.sp_client_id,
                cache_path=cfg.sp_token_cache,
            )
            print("  Writeback    : enabled (column F)")

        # Process each section
        out_dir = Path(args.out_dir)
        all_results: dict[str, list[ProcessResult]] = {}

        try:
            for slug, url_rows in sections.items():
                if not url_rows:
                    print(
                        f"\n  ── {SECTION_LABELS.get(slug, slug).upper()} — no URLs found, skipping"
                    )
                    continue
                label_str = SECTION_LABELS.get(slug, slug)
                results = process_section(
                    label_str,
                    url_rows,
                    fetch_script,
                    args.dry_run,
                    not args.no_validate,
                    wb_ctx=wb_ctx,
                )
                all_results[slug] = results
                out_path = out_dir / f"{sheet_name}-{slug}.txt"
                print()
                write_playlist(out_path, results)
        finally:
            # Close the last workbook session (best-effort)
            if wb_ctx is not None:
                _close_workbook_session(wb_ctx)

        # Write combined failures log
        print()
        failure_path = out_dir / f"{sheet_name}-failures.txt"
        n_failures = write_failures(failure_path, all_results)

        # Summary
        total_processed = sum(len(v) for v in all_results.values())
        total_ok = sum(r.success for v in all_results.values() for r in v)
        print()
        print("  ── Summary ──────────────────────────────────")
        print(f"     Processed : {total_processed}")
        print(f"     Success   : {total_ok}")
        print(f"     Failures  : {n_failures}")
        if wb_ctx and not args.dry_run:
            print(f"     Col F ok  : {wb_ctx.writes_ok}")
            print(f"     Col F err : {wb_ctx.writes_fail}")
        if n_failures:
            print(f"     (see {failure_path})")
        print()


# ── Headless entry point for the webqueue job runner ───────────────────────────


def upcoming_weekend_sheet(today=None) -> tuple[str, "datetime.date", "datetime.date"]:
    """Return (sheet_name, friday, saturday) for the upcoming weekend.

    Per OQ-3: ``friday = today + ((4 - weekday) % 7)`` yields *today* when run
    on a Friday (the imminent weekend); Sat/Sun roll forward to next Friday.
    Sheet name matches ``_SHEET_DATE_RE`` (e.g. ``3.6-3.7``).
    """
    import datetime as _dt

    today = today or _dt.date.today()
    friday = today + _dt.timedelta(days=((4 - today.weekday()) % 7))
    saturday = friday + _dt.timedelta(days=1)
    sheet = f"{friday.month}.{friday.day}-{saturday.month}.{saturday.day}"
    return sheet, friday, saturday


def _make_inprocess_fetch(config):
    """Build a ``run_fetch``-compatible callable backed by the yt-pipe downloader.

    Signature mirrors the original ``run_fetch(url, fetch_script, dry_run)`` and
    returns ``(ok, resolved_manifest_url, note)`` so the rest of the pipeline is
    unchanged. Replaces the powershell/fetch.ps1 subprocess (unavailable in a
    headless service) with an in-process download+upload.
    """
    from ..ytpipe import downloader as _yt

    api_url = f"{config.mediacms_url.rstrip('/')}/api/v1"
    cookies = getattr(config, "fetch_cookies_path", "") or None
    download_dir = (
        str(Path(config.image_dir).parent / "fetch-tmp")
        if getattr(config, "image_dir", None)
        else None
    )

    def _fetch(url: str, fetch_script, dry_run: bool = False):
        if dry_run:
            return True, f"{config.mediacms_url.rstrip('/')}/view?m=DRY_RUN", "dry-run"
        try:
            url2 = _yt.clean_youtube_url(url)
            uploader = _yt.MediaDownloaderToMediaCMS(
                api_url=api_url,
                api_token=config.mediacms_token,
                download_dir=download_dir,
                cookies_file=cookies,
            )
            result = uploader.process_video(url=url2, quality="medium", cleanup=True)
            token = result.get("friendly_token")
            if token:
                manifest = f"{config.mediacms_url.rstrip('/')}/api/v1/media/cytube/{token}.json?format=json"
                return (
                    True,
                    manifest,
                    ("already exists" if result.get("already_exists") else "uploaded"),
                )
            return False, "", result.get("error") or "no token returned"
        except Exception as exc:  # noqa: BLE001 - surface as a per-URL failure
            return False, "", f"{type(exc).__name__}: {exc}"

    return _fetch


def run(params: dict, *, config, progress=None) -> dict:
    """Resolve the upcoming weekend's Channel Z workbook into playlist sections.

    Source precedence:
      1. ``params['workbook_path']`` (one-off local override, e.g. for tests)
      2. SharePoint via Microsoft Graph when ``config.fetchurls.sharepoint_*``
         are set (token acquired *silently* from the pre-seeded MSAL cache)
      3. ``config.fetchurls.workbook_path`` (local file fallback)

    When the source is SharePoint and ``writeback`` is on (and not a dry run),
    resolved dropsugar URLs are written back to column F via the Graph Excel
    API. Off-site URLs are downloaded via the in-process yt-pipe downloader.

    Returns counts plus, per section slug, the resolved ``cm:`` lines and the
    human label ("Friday Night", etc.) so the job wrapper can import each into a
    fixed, well-known saved playlist.
    """
    _check_openpyxl()
    import openpyxl as _oxl  # noqa: F811 - explicit local import for clarity

    section = params.get("section") or "all"
    dry_run = bool(params.get("dry_run", False))
    validate = bool(params.get("validate", True))
    writeback = bool(params.get("writeback", True))

    def _emit(detail):
        if progress:
            progress(detail)

    cfg = getattr(config, "fetchurls", None)
    local_override = params.get("workbook_path") or ""
    sp_tenant = getattr(cfg, "sharepoint_tenant_id", "") if cfg else ""
    sp_client = getattr(cfg, "sharepoint_client_id", "") if cfg else ""
    sp_share = getattr(cfg, "sharepoint_sharing_url", "") if cfg else ""
    sp_cache = getattr(cfg, "token_cache_path", "") if cfg else ""
    cfg_local = getattr(cfg, "workbook_path", "") if cfg else ""

    use_sharepoint = bool(sp_tenant and sp_client and sp_share) and not local_override

    wb_ctx = None
    drive_id = item_id = ""
    graph_token = ""

    if use_sharepoint:
        _emit({"phase": "auth", "source": "sharepoint"})
        graph_token = acquire_graph_token_silent(sp_tenant, sp_client, sp_cache)
        if not graph_token:
            raise RuntimeError(
                "SharePoint authentication unavailable: no valid token in the MSAL "
                "cache. Run a one-time sign-in on the server: "
                "python -m kryten_webqueue.jobs.fetchurls_auth"
            )
        _emit({"phase": "download", "source": "sharepoint"})
        try:
            wb_bytes, drive_id, item_id = download_sharepoint_xlsx(
                graph_token, sp_share
            )
        except SystemExit as exc:  # the vendored reader uses sys.exit on failure
            raise RuntimeError(f"SharePoint download failed: {exc}") from exc
    else:
        workbook_path = local_override or cfg_local
        if not workbook_path:
            raise RuntimeError(
                "fetchurls needs a workbook source: configure "
                "fetchurls.sharepoint_* (recommended) or fetchurls.workbook_path."
            )
        wb_file = Path(workbook_path)
        if not wb_file.exists():
            raise RuntimeError(f"Workbook not found: {wb_file}")
        wb_bytes = wb_file.read_bytes()

    sheet_name, friday, saturday = upcoming_weekend_sheet()
    wb_peek = _oxl.load_workbook(io.BytesIO(wb_bytes), read_only=True)
    all_sheets = wb_peek.sheetnames
    wb_peek.close()
    if sheet_name not in all_sheets:
        # Suggest only date-format weekend sheets (ignore Sheet1/Played Movies/etc).
        weekend_sheets = [s for s in all_sheets if _SHEET_DATE_RE.match(s.strip())]
        available = (
            ", ".join(weekend_sheets) if weekend_sheets else ", ".join(all_sheets)
        )
        raise RuntimeError(
            f"This weekend's worksheet '{sheet_name}' was not found in the workbook. "
            f"Add a sheet named '{sheet_name}' (Friday.date-Saturday.date), or check "
            f"the sheet name matches. Available weekend sheets: {available}"
        )

    _emit({"phase": "parsing", "sheet": sheet_name})
    sections = parse_excel_sections(wb_bytes, sheet_name)
    if section and section != "all":
        sections = {k: v for k, v in sections.items() if k == section}

    # Build the writeback context (SharePoint only, non-dry-run, writeback on).
    if use_sharepoint and writeback and not dry_run:
        wb_ctx = WritebackContext(
            token=graph_token,
            drive_id=drive_id,
            item_id=item_id,
            sheet_name=sheet_name,
            dry_run=False,
            tenant_id=sp_tenant,
            client_id=sp_client,
            cache_path=sp_cache,
        )

    # Swap in the in-process fetch (no fetch.ps1 subprocess in a service).
    global run_fetch  # noqa: PLW0603
    original_run_fetch = run_fetch
    run_fetch = _make_inprocess_fetch(config)

    out_dir = (
        Path(config.image_dir).parent / "fetchurls-playlists"
        if getattr(config, "image_dir", None)
        else Path("fetchurls-playlists")
    )
    out_dir.mkdir(parents=True, exist_ok=True)

    resolved = 0
    downloaded = 0
    failures = 0
    section_lines: dict[str, list[str]] = {}
    section_labels: dict[str, str] = {}
    all_results: dict[str, list[ProcessResult]] = {}
    # Per-section {resolved, failed} counts and a flat list of failed rows for
    # actionable diagnostics (surfaced in the job log + job_runs detail).
    section_summary: dict[str, dict] = {}
    failure_details: list[dict] = []

    try:
        for slug, url_rows in sections.items():
            if not url_rows:
                continue
            label = SECTION_LABELS.get(slug, slug)
            section_labels[slug] = label
            results = process_section(
                label,
                url_rows,
                Path("."),
                dry_run,
                validate,
                wb_ctx=wb_ctx,
            )
            all_results[slug] = results
            write_playlist(out_dir / f"{sheet_name}-{slug}.txt", results)
            lines = []
            sec_resolved = 0
            sec_failed = 0
            for r in results:
                if r.success:
                    resolved += 1
                    sec_resolved += 1
                    if r.resolved_url != r.original_url:
                        downloaded += 1
                    token = _extract_manifest_token(r.resolved_url)
                    lines.append(f"cm:{token}" if token else r.resolved_url)
                else:
                    failures += 1
                    sec_failed += 1
                    failure_details.append(
                        {
                            "section": label,
                            "row": r.row_number,
                            "url": r.original_url,
                            "note": r.note,
                        }
                    )
            section_lines[slug] = lines
            section_summary[label] = {"resolved": sec_resolved, "failed": sec_failed}
        write_failures(out_dir / f"{sheet_name}-failures.txt", all_results)
    finally:
        run_fetch = original_run_fetch
        if wb_ctx is not None:
            try:
                _close_workbook_session(wb_ctx)
            except Exception:  # noqa: BLE001 - best-effort cleanup
                pass

    writeback_stats = None
    if wb_ctx is not None:
        writeback_stats = {"ok": wb_ctx.writes_ok, "failed": wb_ctx.writes_fail}

    # Sync played movies to the Played Movies worksheet (same workbook, idempotent).
    # Skipped when using a local workbook (no Graph credentials to write back with).
    played_movies_result: dict | None = None
    if use_sharepoint:
        from .playedmovies import sync_played_movies as _sync_played

        try:
            played_movies_result = _sync_played(
                wb_bytes,
                sheet_name,
                friday,
                saturday,
                graph_token=graph_token,
                drive_id=drive_id,
                item_id=item_id,
                dry_run=dry_run,
                progress=_emit,
            )
        except (
            Exception
        ) as _pm_exc:  # noqa: BLE001 - non-fatal; keep fetchurls result intact
            logger.warning("played_movies sync failed: %s", _pm_exc, exc_info=True)

    _emit(
        {
            "phase": "done",
            "sheet": sheet_name,
            "resolved": resolved,
            "failures": failures,
        }
    )
    return {
        "sheet": sheet_name,
        "source": "sharepoint" if use_sharepoint else "local",
        "resolved": resolved,
        "downloaded": downloaded,
        "failures": failures,
        "writeback": writeback_stats,
        "section_lines": section_lines,
        "section_labels": section_labels,
        "section_summary": section_summary,
        "failure_details": failure_details,
        "imported_playlists": [],  # filled in by the async job wrapper
        "played_movies": played_movies_result,
        "dry_run": dry_run,
    }


def _extract_manifest_token(url: str) -> str | None:
    """Pull the friendly_token out of a dropsugar manifest/view URL."""
    m = re.search(r"/media/cytube/([^./]+)\.json", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&]m=([^&]+)", url)
    if m:
        return m.group(1)
    return None


if __name__ == "__main__":
    main()
