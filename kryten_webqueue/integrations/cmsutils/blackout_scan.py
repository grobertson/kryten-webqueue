"""Scan the Channel Z workbook for upcoming-weekend dropsugar URLs to black out.

A *blackout* hides a rehosted catalog item from regular users until the end of
the weekend **following** the one it is scheduled for, so replays stay hidden
through that week. This module only *reads* the workbook (SharePoint or a local
``.xlsx``) and resolves dropsugar URLs in columns **E** and **F** to MediaCMS
friendly_tokens; it downloads nothing. The async job wrapper
(:func:`kryten_webqueue.jobs.tasks.catalog_blackout_job`) persists the result.

Follows the vendored-tool contract: ``run(params, *, config, progress=None)``.
"""

from __future__ import annotations

import datetime as _dt
import io
import logging
from typing import Optional

from ...playlists.importer import extract_dropsugar_token
from .fetchurls import (
    _SHEET_DATE_RE,
    acquire_graph_token_silent,
    download_sharepoint_xlsx,
    is_dropsugar,
)

logger = logging.getLogger(__name__)

# Column indices (0-based): E = source URLs, F = resolved dropsugar URLs.
_COL_E = 4
_COL_F = 5

# Days from the sheet's Friday to the blackout expiry. The item airs that
# weekend (Fri..Sun = Friday+2); it stays hidden through the *following* weekend
# (Sun = Friday+9), lifting the Monday after (Friday+10, 00:00).
_BLACKOUT_DAYS = 10


def _sheet_weekend_dates(
    name: str, today: _dt.date
) -> Optional[tuple[_dt.date, _dt.date]]:
    """Return (friday, sunday) for a ``M.D-M.D`` sheet name, or None.

    The sheet name carries no year, so the year is inferred as the nearest
    occurrence to ``today`` (handles year-boundary sheets like ``12.26-12.27``).
    """
    m = _SHEET_DATE_RE.match(name.strip())
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    best: Optional[_dt.date] = None
    for year in (today.year - 1, today.year, today.year + 1):
        try:
            cand = _dt.date(year, month, day)
        except ValueError:
            continue
        if best is None or abs((cand - today).days) < abs((best - today).days):
            best = cand
    if best is None:
        return None
    return best, best + _dt.timedelta(days=2)


def _expiry_for_friday(friday: _dt.date) -> str:
    """Blackout expiry (SQLite-comparable ``YYYY-MM-DD HH:MM:SS``) for a weekend."""
    monday = friday + _dt.timedelta(days=_BLACKOUT_DAYS)
    return _dt.datetime.combine(monday, _dt.time.min).isoformat(sep=" ")


def scan_workbook_blackouts(wb_bytes: bytes, today: _dt.date) -> tuple[list[dict], int]:
    """Return (blackouts, sheets_scanned) for current/future weekend sheets.

    ``blackouts`` is a list of ``{"token", "reason", "expires_at"}``; a token
    seen across multiple sheets keeps the latest expiry.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    try:
        found: dict[str, dict] = {}
        sheets_scanned = 0
        for name in wb.sheetnames:
            dates = _sheet_weekend_dates(name, today)
            if not dates:
                continue
            friday, sunday = dates
            if sunday < today:  # weekend already fully past
                continue
            sheets_scanned += 1
            expires_at = _expiry_for_friday(friday)
            reason = f"weekend:{name.strip()}"
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                for idx in (_COL_E, _COL_F):
                    if len(row) <= idx or row[idx] is None:
                        continue
                    val = str(row[idx]).strip()
                    if not val or not is_dropsugar(val):
                        continue
                    token = extract_dropsugar_token(val)
                    if not token:
                        continue
                    prev = found.get(token)
                    if prev is None or expires_at > prev["expires_at"]:
                        found[token] = {
                            "token": token,
                            "reason": reason,
                            "expires_at": expires_at,
                        }
        return list(found.values()), sheets_scanned
    finally:
        wb.close()


def run(params: dict, *, config, progress=None) -> dict:
    """Resolve upcoming-weekend dropsugar URLs into blackout entries.

    Source precedence mirrors :func:`fetchurls.run`:
      1. ``params['workbook_path']`` (local override)
      2. SharePoint via Graph when ``config.fetchurls.sharepoint_*`` are set
      3. ``config.fetchurls.workbook_path`` (local fallback)

    Returns ``{"source", "sheets_scanned", "blackouts": [...]}``. Never writes
    the workbook or downloads media.
    """
    from pathlib import Path

    def _emit(detail: dict) -> None:
        if progress:
            progress(detail)

    cfg = getattr(config, "fetchurls", None)
    local_override = params.get("workbook_path") or ""
    sp_tenant = getattr(cfg, "sharepoint_tenant_id", "") if cfg else ""
    sp_client = getattr(cfg, "sharepoint_client_id", "") if cfg else ""
    sp_share = getattr(cfg, "sharepoint_sharing_url", "") if cfg else ""
    sp_cache = getattr(cfg, "token_cache_path", "") if cfg else ""
    cfg_local = getattr(cfg, "workbook_path", "") if cfg else ""

    use_sharepoint = bool(sp_tenant and sp_client and sp_share) and not local_override

    if use_sharepoint:
        _emit({"phase": "auth", "source": "sharepoint"})
        token = acquire_graph_token_silent(sp_tenant, sp_client, sp_cache)
        if not token:
            raise RuntimeError(
                "SharePoint authentication unavailable: no valid token in the MSAL "
                "cache. Run a one-time sign-in on the server: "
                "python -m kryten_webqueue.jobs.fetchurls_auth"
            )
        _emit({"phase": "download", "source": "sharepoint"})
        wb_bytes, _drive_id, _item_id = download_sharepoint_xlsx(token, sp_share)
        source = "sharepoint"
    else:
        workbook_path = local_override or cfg_local
        if not workbook_path:
            raise RuntimeError(
                "catalog_blackout needs a workbook source: configure "
                "fetchurls.sharepoint_* (recommended) or fetchurls.workbook_path."
            )
        wb_file = Path(workbook_path)
        if not wb_file.exists():
            raise RuntimeError(f"Workbook not found: {wb_file}")
        wb_bytes = wb_file.read_bytes()
        source = "local"

    _emit({"phase": "scanning"})
    blackouts, sheets_scanned = scan_workbook_blackouts(wb_bytes, _dt.date.today())
    _emit(
        {
            "phase": "done",
            "sheets_scanned": sheets_scanned,
            "blackouts": len(blackouts),
        }
    )
    return {
        "source": source,
        "sheets_scanned": sheets_scanned,
        "blackouts": blackouts,
    }
