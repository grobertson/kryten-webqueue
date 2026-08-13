"""Played Movies worksheet sync.

Extracts movies from a weekend schedule sheet (col B = type tag, col C = title)
and records them idempotently to the "Played Movies" sheet in the same workbook.
Used by the fetchurls job after resolving URL sections; may also run standalone.

Column layout in the Played Movies sheet:
  A — movie title (text)
  B — air date as an Excel date serial number (integer; cell is date-formatted)
"""

from __future__ import annotations

import datetime
import io
import json
import logging
import time

import requests

logger = logging.getLogger(__name__)

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
PLAYED_SHEET = "Played Movies"
MOVIE_SECTIONS = {"friday", "saturday-night"}
WRITEBACK_TIMEOUT = 30
WRITEBACK_DELAY = 0.5  # seconds between Graph API writes

# Section header substrings → slug (order matters: more specific first)
_SECTION_MAP = {
    "friday": "friday",
    "saturday morning": "saturday-morning",
    "saturday": "saturday-night",
}


def _excel_date_serial(d: datetime.date) -> int:
    """Convert a Python date to an Excel date serial (days since 1899-12-30)."""
    return (d - datetime.date(1899, 12, 30)).days


def _classify_section(cell_value: str) -> str | None:
    if not cell_value:
        return None
    v = cell_value.strip().lower()
    for keyword, slug in _SECTION_MAP.items():
        if keyword in v:
            return slug
    return None


def _is_movie_tag(value: str) -> bool:
    v = value.strip().lower()
    return v == "movie" or v == "movie event"


def extract_movies(wb_bytes: bytes, sheet_name: str) -> list[tuple[str, str]]:
    """Return [(title, section_slug), ...] for movie rows in the weekend sheet.

    Reads col A (section header), col B (type tag), col C (title).
    Only returns rows where col B is "Movie" or "Movie Event" in Friday or
    Saturday Night sections.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for played-movies sync")

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]
    current_section: str | None = None
    movies: list[tuple[str, str]] = []

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0]).strip() if row[0] is not None else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        slug = _classify_section(col_a)
        if slug:
            current_section = slug
            continue

        if current_section in MOVIE_SECTIONS and _is_movie_tag(col_b) and col_c:
            movies.append((col_c, current_section))

    wb.close()
    return movies


def read_played_movie_titles(wb_bytes: bytes) -> set[str]:
    """Return the lowercased title set from column A of the Played Movies sheet."""
    try:
        import openpyxl
    except ImportError:
        return set()

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    if PLAYED_SHEET not in wb.sheetnames:
        wb.close()
        return set()

    ws = wb[PLAYED_SHEET]
    titles: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        val = str(row[0]).strip() if row[0] is not None else ""
        if val:
            titles.add(val.lower())
    wb.close()
    return titles


def last_played_row(wb_bytes: bytes) -> int:
    """Return the 1-based index of the last non-empty row in Played Movies (0 if absent)."""
    try:
        import openpyxl
    except ImportError:
        return 0

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    if PLAYED_SHEET not in wb.sheetnames:
        wb.close()
        return 0

    ws = wb[PLAYED_SHEET]
    last = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        val = str(row[0]).strip() if row[0] is not None else ""
        if val:
            last = i
    wb.close()
    return last


def _write_rows(
    token: str,
    drive_id: str,
    item_id: str,
    new_movies: list[tuple[str, int]],
    start_row: int,
    *,
    progress=None,
) -> tuple[int, int]:
    """Write (title, excel_date_serial) pairs to the Played Movies sheet via Graph.

    Returns (ok_count, fail_count).
    """
    wb_base = f"{GRAPH_BASE}/drives/{drive_id}/items/{item_id}/workbook"
    sheet_enc = PLAYED_SHEET.replace("'", "''")
    ok = fail = 0

    for i, (title, date_serial) in enumerate(new_movies):
        row_num = start_row + i
        if progress:
            progress({"phase": "played_movies_write", "row": row_num, "title": title})

        sess_resp = requests.post(
            f"{wb_base}/createSession",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            data=json.dumps({"persistChanges": True}),
            timeout=WRITEBACK_TIMEOUT,
        )
        sid = ""
        if sess_resp.status_code in (200, 201):
            sid = sess_resp.json().get("id", "")

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }
        if sid:
            headers["workbook-session-id"] = sid

        url = f"{wb_base}/worksheets('{sheet_enc}')/range(address='A{row_num}:B{row_num}')"
        try:
            resp = requests.patch(
                url,
                headers=headers,
                data=json.dumps({"values": [[title, date_serial]]}),
                timeout=WRITEBACK_TIMEOUT,
            )
            vals = resp.json().get("values", [[]])
            wrote_title = vals[0][0] if vals and len(vals[0]) > 0 else ""
            if resp.status_code in (200, 204) and wrote_title == title:
                logger.info("played_movies: wrote row %d: %s", row_num, title)
                ok += 1
            else:
                logger.warning(
                    "played_movies: row %d write failed (HTTP %s, got %r)",
                    row_num,
                    resp.status_code,
                    wrote_title,
                )
                fail += 1
        except requests.RequestException as exc:
            logger.warning("played_movies: row %d request error: %s", row_num, exc)
            fail += 1
        finally:
            if sid:
                try:
                    requests.post(
                        f"{wb_base}/closeSession",
                        headers={
                            "Authorization": f"Bearer {token}",
                            "Content-Type": "application/json",
                            "workbook-session-id": sid,
                        },
                        data="{}",
                        timeout=WRITEBACK_TIMEOUT,
                    )
                except requests.RequestException:
                    pass

        time.sleep(WRITEBACK_DELAY)

    return ok, fail


def sync_played_movies(
    wb_bytes: bytes,
    sheet_name: str,
    friday: datetime.date,
    saturday: datetime.date,
    *,
    graph_token: str,
    drive_id: str,
    item_id: str,
    dry_run: bool = False,
    progress=None,
) -> dict:
    """Idempotently sync played movies from the weekend sheet to Played Movies.

    Returns {"found": n, "added": n, "skipped": n, "failed": n, "dry_run": bool}.
    """
    if progress:
        progress({"phase": "played_movies_scan"})

    try:
        movies = extract_movies(wb_bytes, sheet_name)
    except RuntimeError as exc:
        logger.warning("played_movies: skipping — %s", exc)
        return {"found": 0, "added": 0, "skipped": 0, "failed": 0, "dry_run": dry_run}

    if not movies:
        return {"found": 0, "added": 0, "skipped": 0, "failed": 0, "dry_run": dry_run}

    existing = read_played_movie_titles(wb_bytes)
    fri_serial = _excel_date_serial(friday)
    sat_serial = _excel_date_serial(saturday)

    new_rows: list[tuple[str, int]] = []
    skipped = 0
    for title, section in movies:
        if title.lower() in existing:
            skipped += 1
            continue
        serial = fri_serial if section == "friday" else sat_serial
        new_rows.append((title, serial))

    if not new_rows:
        logger.info("played_movies: all %d movie(s) already recorded", len(movies))
        return {
            "found": len(movies),
            "added": 0,
            "skipped": skipped,
            "failed": 0,
            "dry_run": dry_run,
        }

    start_row = last_played_row(wb_bytes) + 1

    if dry_run:
        logger.info(
            "played_movies [dry-run]: would add %d movie(s) starting at row %d",
            len(new_rows),
            start_row,
        )
        return {
            "found": len(movies),
            "added": len(new_rows),
            "skipped": skipped,
            "failed": 0,
            "dry_run": True,
        }

    if progress:
        progress({"phase": "played_movies_write", "count": len(new_rows)})

    ok, fail = _write_rows(
        graph_token, drive_id, item_id, new_rows, start_row, progress=progress
    )
    logger.info(
        "played_movies: added %d/%d, skipped %d, failed %d",
        ok,
        len(new_rows),
        skipped,
        fail,
    )
    return {
        "found": len(movies),
        "added": ok,
        "skipped": skipped,
        "failed": fail,
        "dry_run": False,
    }
